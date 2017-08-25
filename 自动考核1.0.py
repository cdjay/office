import random
import openpyxl

#设置部分
staffSheet=[]
staffSheet.append(['魏小龙'])#动作
staffSheet.append(['周栩'])#音效
staffSheet.append(['王娜'])#3D角色
staffSheet.append(['陈幸','付云丽','梁馨竹','唐敬','李娟','谢佳','吴佳雯','刘欣禹'])#2D角色
staffSheet.append(['唐嘉俊'])#特效
staffSheet.append(['王瀚冰','李娜'])#UI
staffSheet.append([''])#后期
staffSheet.append(['钟放'])#3D场景
staffSheet.append(['金飞','王兴文','唐小雯'])#2D场景
skillSheet=['动作','音效','3D角色','2D角色','特效','UI','后期','3D场景','2D场景']
leaderSheet=["张鑫","NaN","曾敏","赵鑫","NaN","郑星杰","朱易","吴云峰","none"]
global total
global totalss
total=[]
totalss=[]

def staff(name,skill): #生成员工考核成绩表
    s=0#初始分数归0
    #随机生成分数
    while s<70 or s>79:            #合格分数线为70-79分
        s1=random.randint(18, 30)   #任务完成质量
        s2=random.randint(18, 30)   #工时完成率
        s3=random.randint(12, 20)   #协作配合度
        s4=random.randint(12, 20)   #工作态度与纪律
        s=s1+s2+s3+s4               #总分数
    #读写文件
    book = "美术绩效考核表-"+str(name)+".xlsx"
    wb = openpyxl.load_workbook('sample.xlsx')#打开文件
    ws = wb.active#激活文件
    #填充数据
    ws['J6'] = s1
    ws['J7'] = s2
    ws['J8'] = s3
    ws['J9'] = s4
    ws['C3'] = name #姓名
    ws['F3'] = skill #姓名
    ws['H3'] = 168  #工时
    totalss.append(s)#统计名字
    total.append(name)#统计分数
    wb.save(book)# 保存文件
    return  s

def leader(name):#生成组长考核成绩表
    pass

def printotal():#统计总分
    wb = openpyxl.load_workbook('total.xlsx')#打开文件
    ws = wb.active#激活文件
    for i in range(len(total)):
        pos='B'+str(i+2)
        ws[pos] = total[i]#填入姓名
        pos='C'+str(i+2)
        ws[pos] = totalss[i]#填入分数
    wb.save('考核总分.xlsx')# 保存文件

def makefiles():#生成考核表
    for i in range(len(skillSheet)):
    # for i in range(1):
        for j in range(len(staffSheet[i])):
            print('正在生成员工:[%s]的考核表'%(staffSheet[i][j]))
            staff(staffSheet[i][j],skillSheet[i])

#main()
makefiles()
# printotal()















# #组长考核
# for i in range(len(leaderSheet)):
#     print('正在生成组长:[%s]的考核表'%(leaderSheet[i]))
#     print(leader(leaderSheet[i]))