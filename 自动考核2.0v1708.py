import openpyxl # 载入excel操作包
#设置部分
leaderList=['张鑫','曾敏','郑星杰','朱易','吴云峰','赵鑫','刘强'] # 组长表
staffList=['张鑫','魏小龙','周佳欣','周栩','曾敏','王娜','赵鑫','陈幸','付云丽','梁馨竹','唐敬','李娟','谢佳','吴佳雯','刘欣禹','张文军','陈方欣','唐嘉俊','郑星杰','王瀚冰','李娜','朱易','吴云峰','钟放','刘强','金飞','王兴文','唐小雯']# 员工总表
for row in range(4,32): # 逐行读取所有人员数据 column数据从2开始.row数据从4开始.总人数28人.即循环终点为4+28=32
    #读取总表信息.包括姓名,部门,绩效,分数.四个值
    wb = openpyxl.load_workbook('考评总计.xlsx') #打开总表
    ws=wb.get_sheet_by_name('Sheet1') # 打开分表
    staffDep=ws.cell(row=row, column=2).value # 部门
    staffName=ws.cell(row=row, column=3).value # 姓名
    staffSoc=ws.cell(row=row, column=4).value # 总分数
    staffJob=ws.cell(row=row, column=5).value # 绩效时间
    #判断员工是否为组长
    if staffName in leaderList: # 如果员工为组长,则生成组长表格
        wb = openpyxl.load_workbook('leader.xlsx') #打开考核表
        ws=wb.get_sheet_by_name('组长') # 打开分表
        # 生成各项分数
        s2=int(35*(staffSoc/100))
        s3=int(10*(staffSoc/100))
        s4=int(5*(staffSoc/100))
        s5=int(5*(staffSoc/100))
        s6=int(15*(staffSoc/100))
        s1=staffSoc-s6-s2-s3-s4-s5
        #写入表格
        ws['D2']=staffName
        ws['G2']=staffDep+'组长'
        ws['I2']=staffJob
        ws['K5']=s1
        ws['K6']=s2
        ws['K7']=s3
        ws['K8']=s4
        ws['K9']=s5
        ws['K10']=s6
        # 存档到out目录
        wb.save(filename='out\美术绩效考核表-组长-'+staffName+'.xlsx')
    else: # 否则生成员工表格
        wb = openpyxl.load_workbook('staff.xlsx') #打开考核表
        ws=wb.get_sheet_by_name('组员') # 打开分表
        # 生成各项分数
        s1=int(30*(staffSoc/100))
        s2=int(30*(staffSoc/100))
        s3=int(20*(staffSoc/100))
        s4=staffSoc-s1-s2-s3
        #写入表格
        ws['C3']=staffName
        ws['F3']=staffDep
        ws['H3']=staffJob
        ws['J6']=s1
        ws['J7']=s2
        ws['J8']=s3
        ws['J9']=s4
        # 存档到out目录
        wb.save(filename='out\美术绩效考核表-'+staffName+'.xlsx')